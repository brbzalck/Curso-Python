# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Nome para planilha
sheet_name = 'Minha planilha'
# Criando planilha
workbook.create_sheet(sheet_name, 0)
# Criando a planilha com o nome criado
worksheet: Worksheet = workbook[sheet_name]
print(workbook.sheetnames)

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['Joao', 14, 4.5],
    ['Lucas', 13, 4.7],
    ['Maria', 15, 4.4],
    ['Duda', 16, 4.2]
]
# MODELO UTIL
# # Numerando litas internas
# for i, student_row in enumerate(students, start=2):
#     # Numerando itens(colunas) das listas interna
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
